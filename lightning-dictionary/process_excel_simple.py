#!/usr/bin/env python3
"""
Excel to JSON converter for Lightning Dictionary
Processes wordFrequency.xlsx to create optimized dictionary data
"""

import json
import os
from pathlib import Path

def process_dictionary():
    """
    Process wordFrequency.xlsx file and convert to JSON
    Designed to be simple and easy to understand for beginners
    """
    
    # Configuration
    excel_file = "data/wordFrequency.xlsx"
    output_dir = "data/processed"
    sheet_name = "1 lemmas"  # As specified in CLAUDE.md
    word_column_index = 1  # Column B (0=A, 1=B, 2=C, etc.)
    max_words = 10000  # Top 10,000 words for Phase 1
    
    print("Lightning Dictionary - Data Processing")
    print("=" * 40)
    print(f"Input file: {excel_file}")
    print(f"Output directory: {output_dir}")
    print(f"Maximum words: {max_words}")
    print()
    
    # Step 1: Check dependencies
    try:
        import openpyxl
        print("✓ openpyxl is installed")
    except ImportError:
        print("✗ Error: openpyxl is not installed")
        print("\nTo install it, run:")
        print("  pip install openpyxl")
        return False
    
    # Step 2: Check if Excel file exists
    if not os.path.exists(excel_file):
        print(f"✗ Error: Excel file not found at: {excel_file}")
        print("\nPlease ensure the file exists at the specified location.")
        print("Expected structure:")
        print("  lightning-dictionary/")
        print("  ├── data/")
        print("  │   └── wordFrequency.xlsx")
        print("  └── process_dictionary.py (this script)")
        return False
    
    print(f"✓ Found Excel file")
    
    # Step 3: Create output directory
    os.makedirs(output_dir, exist_ok=True)
    print(f"✓ Output directory ready: {output_dir}")
    
    # Step 4: Load and process Excel file
    try:
        print("\nLoading Excel file...")
        workbook = openpyxl.load_workbook(excel_file, read_only=True)
        
        # Check available sheets
        print(f"Available sheets: {workbook.sheetnames}")
        
        # Select the correct sheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"✓ Using sheet: '{sheet_name}'")
        else:
            # Fallback to first sheet if specified sheet not found
            sheet = workbook.active
            print(f"⚠ Sheet '{sheet_name}' not found, using: '{sheet.title}'")
        
        # Process the data
        print(f"\nProcessing words from column {chr(65 + word_column_index)} (column {word_column_index + 1})...")
        
        words_dict = {}
        words_processed = 0
        errors = 0
        
        # Iterate through rows (skip header row)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if words_processed >= max_words:
                break
            
            # Skip empty rows
            if not row or not any(row):
                continue
            
            try:
                # Extract data from columns
                # Adjust these indices based on your Excel structure
                rank = row[0] if row[0] else words_processed + 1  # Column A
                word = row[word_column_index] if len(row) > word_column_index else None  # Column B
                pos = row[2] if len(row) > 2 and row[2] else 'n/a'  # Column C (part of speech)
                frequency = row[3] if len(row) > 3 and row[3] else 0  # Column D
                
                # Skip if no word found
                if not word:
                    continue
                
                # Clean and normalize the word
                word = str(word).strip().lower()
                
                # Skip empty words
                if not word:
                    continue
                
                # Create word entry
                word_entry = {
                    'rank': int(rank),
                    'pos': str(pos).strip(),
                    'frequency': int(frequency) if frequency else 0,
                    # These will be populated later when integrating with dictionary APIs
                    'definitions': [],
                    'pronunciation': '',
                    'examples': []
                }
                
                words_dict[word] = word_entry
                words_processed += 1
                
                # Show progress
                if words_processed % 1000 == 0:
                    print(f"  Processed {words_processed} words...")
                
            except Exception as e:
                errors += 1
                if errors <= 5:  # Only show first 5 errors
                    print(f"  ⚠ Error on row {row_num}: {str(e)}")
                continue
        
        workbook.close()
        
        print(f"\n✓ Processing complete:")
        print(f"  - Total words processed: {words_processed}")
        print(f"  - Errors encountered: {errors}")
        
    except Exception as e:
        print(f"\n✗ Error loading Excel file: {str(e)}")
        return False
    
    # Step 5: Add metadata and structure
    output_data = {
        'version': '1.0',
        'metadata': {
            'total_words': len(words_dict),
            'source_file': excel_file,
            'generated_date': str(Path.cwd()),
            'description': 'Lightning Dictionary - Core vocabulary data'
        },
        'statistics': {
            'parts_of_speech': {},
            'frequency_ranges': {
                'top_100': 0,
                'top_1000': 0,
                'top_5000': 0,
                'rest': 0
            }
        },
        'words': words_dict
    }
    
    # Calculate statistics
    for word, data in words_dict.items():
        # Count parts of speech
        pos = data['pos']
        if pos in output_data['statistics']['parts_of_speech']:
            output_data['statistics']['parts_of_speech'][pos] += 1
        else:
            output_data['statistics']['parts_of_speech'][pos] = 1
        
        # Count frequency ranges
        rank = data['rank']
        if rank <= 100:
            output_data['statistics']['frequency_ranges']['top_100'] += 1
        elif rank <= 1000:
            output_data['statistics']['frequency_ranges']['top_1000'] += 1
        elif rank <= 5000:
            output_data['statistics']['frequency_ranges']['top_5000'] += 1
        else:
            output_data['statistics']['frequency_ranges']['rest'] += 1
    
    # Step 6: Save to JSON files
    print("\nSaving JSON files...")
    
    # Pretty-printed version for development
    output_file = os.path.join(output_dir, 'dictionary.json')
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        
        file_size = os.path.getsize(output_file) / (1024 * 1024)  # Convert to MB
        print(f"✓ Development version saved: {output_file}")
        print(f"  Size: {file_size:.2f} MB")
    except Exception as e:
        print(f"✗ Error saving development JSON: {str(e)}")
        return False
    
    # Minified version for production
    minified_file = os.path.join(output_dir, 'dictionary.min.json')
    try:
        with open(minified_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, separators=(',', ':'), ensure_ascii=False)
        
        min_size = os.path.getsize(minified_file) / (1024 * 1024)
        print(f"✓ Production version saved: {minified_file}")
        print(f"  Size: {min_size:.2f} MB (saved {file_size - min_size:.2f} MB)")
    except Exception as e:
        print(f"✗ Error saving minified JSON: {str(e)}")
        return False
    
    # Step 7: Display summary
    print("\n" + "=" * 40)
    print("SUMMARY")
    print("=" * 40)
    print(f"Total words: {output_data['metadata']['total_words']}")
    print("\nParts of speech distribution:")
    for pos, count in sorted(output_data['statistics']['parts_of_speech'].items(), 
                            key=lambda x: x[1], reverse=True)[:5]:
        print(f"  {pos}: {count} words")
    
    print("\nFrequency distribution:")
    for range_name, count in output_data['statistics']['frequency_ranges'].items():
        print(f"  {range_name}: {count} words")
    
    print("\nSample words (top 10):")
    sample_words = sorted(words_dict.items(), key=lambda x: x[1]['rank'])[:10]
    for word, data in sample_words:
        print(f"  #{data['rank']}: {word} ({data['pos']}) - freq: {data['frequency']:,}")
    
    print("\n✅ Data processing complete!")
    print(f"Your dictionary data is ready at: {output_file}")
    
    return True


def validate_json_output():
    """
    Validate the generated JSON file
    """
    json_file = "data/processed/dictionary.json"
    
    if not os.path.exists(json_file):
        print("✗ JSON file not found. Run process_dictionary() first.")
        return False
    
    print("\nValidating JSON output...")
    
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Check structure
        required_keys = ['version', 'metadata', 'statistics', 'words']
        for key in required_keys:
            if key not in data:
                print(f"✗ Missing required key: {key}")
                return False
        
        # Check words
        words = data['words']
        if not words:
            print("✗ No words found in dictionary")
            return False
        
        # Check a sample word
        sample_word = next(iter(words.values()))
        required_word_keys = ['rank', 'pos', 'frequency']
        for key in required_word_keys:
            if key not in sample_word:
                print(f"✗ Words missing required key: {key}")
                return False
        
        print("✓ JSON validation passed")
        print(f"  Total words: {len(words)}")
        print(f"  File is ready for use in the application")
        
        return True
        
    except json.JSONDecodeError as e:
        print(f"✗ Invalid JSON: {str(e)}")
        return False
    except Exception as e:
        print(f"✗ Validation error: {str(e)}")
        return False


if __name__ == "__main__":
    # Run the processing
    success = process_dictionary()
    
    # If successful, validate the output
    if success:
        validate_json_output()
    else:
        print("\n⚠ Processing failed. Please check the errors above.")